"""Monthly todo list tracker with Excel persistence."""

from __future__ import annotations

import shutil
import tkinter as tk
from tkinter import ttk
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional
from uuid import uuid4
from tkinter import messagebox, simpledialog

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_FILE = Path(__file__).with_name("MyToDoList.xlsx")
SYNC_FILE = Path(r"C:\Users\kumar\OneDrive\Desktop\MyToDoList.xlsx")
SHEET_NAME = "Todo List"
TEMPLATE_SHEET_NAME = "TaskSubTask"
HEADERS = (
	"ID",
	"List",
	"Category",
	"Task",
	"Frequency",
	"Bill for the Month",
	"Month Key",
	"Created Date",
	"Completed Date",
	"Status",
)
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"
COMPLETION_DATE_FORMAT = "%Y-%b-%d"
PENDING_STATUS = "Pending"
DONE_STATUS = "Done"
STANDARD_TASK_TEMPLATES = (
	"Monthly report",
	"Monthly budget review",
	"Monthly backup",
	"Monthly planning",
	"Monthly system check",
)
MONTH_OPTIONS = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")


@dataclass
class Task:
	"""A task and the timestamps that describe its lifecycle."""

	id: str
	title: str
	created_at: str
	completed_at: Optional[str] = None
	month: Optional[str] = None
	bill_for_month: Optional[str] = None
	list_name: Optional[str] = None
	category: Optional[str] = None
	frequency: str = "Monthly"

	@property
	def is_done(self) -> bool:
		return self.completed_at is not None


def now() -> str:
	return datetime.now().strftime(DATE_FORMAT)


def month_key(value: str) -> str:
	"""Return the month portion of a stored timestamp."""
	return value[:7]


def task_month(task: Task) -> str:
	return task.month or month_key(task.created_at)


def bill_month(task: Task) -> str:
	if task.bill_for_month:
		return task.bill_for_month
	return datetime.strptime(task_month(task), "%Y-%m").strftime("%b").upper()


def month_key_for_bill(bill_for_month: str) -> str:
	month_number = MONTH_OPTIONS.index(bill_for_month) + 1
	return f"{datetime.now().year:04d}-{month_number:02d}"


def next_month(selected_month: str) -> str:
	parsed_month = datetime.strptime(selected_month, "%Y-%m")
	if parsed_month.month == 12:
		return f"{parsed_month.year + 1}-01"
	return f"{parsed_month.year:04d}-{parsed_month.month + 1:02d}"


def add_months(selected_month: str, months: int) -> str:
	parsed_month = datetime.strptime(selected_month, "%Y-%m")
	total_months = parsed_month.year * 12 + parsed_month.month - 1 + months
	return f"{total_months // 12:04d}-{total_months % 12 + 1:02d}"


def normalize_frequency(value: Optional[str]) -> str:
	return "Quarterly" if str(value or "Monthly").strip().lower() in {"quaterly", "quarterly"} else "Monthly"


def next_task_month(task: Task) -> str:
	return add_months(task_month(task), 3 if normalize_frequency(task.frequency) == "Quarterly" else 1)


def sync_workbook() -> None:
	"""Keep both workbook locations synchronized using the newest file."""
	# Select the newest workbook, then copy it to both locations so either path has the latest data.
	try:
		local_exists = DATA_FILE.exists()
		remote_exists = SYNC_FILE.exists()
		if not local_exists and not remote_exists:
			return
		if not local_exists:
			DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
			shutil.copy2(SYNC_FILE, DATA_FILE)
			return
		if not remote_exists:
			SYNC_FILE.parent.mkdir(parents=True, exist_ok=True)
			shutil.copy2(DATA_FILE, SYNC_FILE)
			return

		source = DATA_FILE if DATA_FILE.stat().st_mtime >= SYNC_FILE.stat().st_mtime else SYNC_FILE
		for destination in (DATA_FILE, SYNC_FILE):
			if destination != source:
				shutil.copy2(source, destination)
	except OSError as error:
		print(f"Could not synchronize MyToDoList.xlsx: {error}")


def load_tasks() -> list[Task]:
	# Synchronize before reading so tasks and TaskSubTask changes from either location are available.
	sync_workbook()
	if not DATA_FILE.exists():
		return []

	try:
		workbook = load_workbook(DATA_FILE, read_only=True, data_only=True)
		if SHEET_NAME not in workbook.sheetnames:
			return []

		worksheet = workbook[SHEET_NAME]
		header = [cell.value for cell in worksheet[1]]
		column = {name: header.index(name) for name in header if name}
		tasks = []
		for row in worksheet.iter_rows(min_row=2, values_only=True):
			if not row[column["ID"]] or not row[column["Task"]] or not row[column["Created Date"]]:
				continue
			tasks.append(
				Task(
					id=str(row[column["ID"]]),
					title=str(row[column["Task"]]),
					created_at=str(row[column["Created Date"]]),
					frequency=(
						normalize_frequency(row[column["Frequency"]])
						if "Frequency" in column and row[column["Frequency"]]
						else "Monthly"
					),
					completed_at=(
						str(row[column["Completed Date"]])
						if "Completed Date" in column and row[column["Completed Date"]]
						else None
					),
					month=(
						str(row[column["Month Key"]])
						if "Month Key" in column and row[column["Month Key"]]
						else str(row[column["Month"]])
						if "Month" in column and row[column["Month"]]
						else None
					),
					bill_for_month=(
						str(row[column["Bill for the Month"]]).upper()
						if "Bill for the Month" in column and row[column["Bill for the Month"]]
						else None
					),
					list_name=(
						str(row[column["List"]])
						if "List" in column and row[column["List"]]
						else None
					),
					category=(
						str(row[column["Category"]])
						if "Category" in column and row[column["Category"]]
						else None
					),
				)
			)
		return tasks
	except (KeyError, OSError, TypeError, ValueError) as error:
		print(f"Could not read {DATA_FILE.name}: {error}")
		return []
	finally:
		if "workbook" in locals():
			workbook.close()


def save_tasks(tasks: list[Task]) -> None:
	# Rebuild only the generated Todo List sheet while preserving the TaskSubTask reference sheet.
	sync_workbook()
	if DATA_FILE.exists():
		workbook = load_workbook(DATA_FILE)
		if SHEET_NAME in workbook.sheetnames:
			del workbook[SHEET_NAME]
		worksheet = workbook.create_sheet(SHEET_NAME, 0)
	else:
		workbook = Workbook()
		worksheet = workbook.active
	worksheet.title = SHEET_NAME
	worksheet.append(HEADERS)

	header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
	completed_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
	for cell in worksheet[1]:
		cell.font = Font(color="FFFFFF", bold=True)
		cell.fill = header_fill

	for task in tasks:
		# Store both the visible billing month and the hidden year-aware key used for scheduling.
		worksheet.append(
			[
				task.id,
				task.list_name or "",
				task.category or "",
				task.title,
				normalize_frequency(task.frequency),
				bill_month(task),
				task_month(task),
				task.created_at,
				task.completed_at or "",
				DONE_STATUS if task.is_done else PENDING_STATUS,
			]
		)
		if task.is_done:
			for cell in worksheet[worksheet.max_row]:
				cell.fill = completed_fill

	worksheet.freeze_panes = "A2"
	worksheet.auto_filter.ref = worksheet.dimensions
	status_validation = DataValidation(type="list", formula1='"Pending,Done"', allow_blank=False)
	worksheet.add_data_validation(status_validation)
	status_validation.add(f"J2:J{max(2, worksheet.max_row)}")
	worksheet.column_dimensions["G"].hidden = True
	column_widths = (34, 18, 24, 42, 14, 20, 14, 22, 22, 12)
	for column_number, width in enumerate(column_widths, start=1):
		worksheet.column_dimensions[get_column_letter(column_number)].width = width

	workbook.save(DATA_FILE)
	workbook.close()
	sync_workbook()


def display_tasks(
	tasks: list[Task],
	selected_month: Optional[str] = None,
	pending_only: bool = False,
) -> None:
	visible_tasks = tasks
	if selected_month:
		visible_tasks = [task for task in tasks if task_month(task) == selected_month]
	if pending_only:
		visible_tasks = [task for task in visible_tasks if not task.is_done]

	if not visible_tasks:
		label = selected_month or "your todo list"
		print(f"\nNo tasks found for {label}.\n")
		return

	visible_tasks.sort(key=lambda task: (task.is_done, task.created_at))
	print(f"\nTodo list{f' for {selected_month}' if selected_month else ''}")
	print("-" * 78)
	for number, task in enumerate(visible_tasks, start=1):
		status = DONE_STATUS if task.is_done else PENDING_STATUS
		completed = task.completed_at or "-"
		location = f" [{task.list_name}]" if task.list_name else ""
		category = f" / {task.category}" if task.category else ""
		print(f"{number}. [{status}] {task.title}{location}{category}")
		print(f"   Frequency: {normalize_frequency(task.frequency)}    Bill for: {bill_month(task)}")
		print(f"   Created: {task.created_at}    Completed: {completed}")
	print()


def add_task(tasks: list[Task], title: str, bill_for_month: str) -> None:
	tasks.append(
		Task(
			id=uuid4().hex,
			title=title,
			created_at=now(),
			month=month_key_for_bill(bill_for_month),
			bill_for_month=bill_for_month,
		)
	)
	save_tasks(tasks)
	print(f"Added: {title}")


def read_template_items(list_name: str) -> list[tuple[str, str, str]]:
	value_column = 1 if list_name == "Pune List" else 4
	frequency_column = 2 if list_name == "Pune List" else 5
	sync_workbook()
	workbook = load_workbook(DATA_FILE, read_only=True, data_only=True)
	try:
		if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
			return []

		worksheet = workbook[TEMPLATE_SHEET_NAME]
		items = []
		category = "General"
		for row in worksheet.iter_rows(min_row=2, values_only=True):
			# A row ending in '-' starts a category; following rows are that category's subtasks.
			value = row[value_column - 1]
			if not value:
				continue
			value = str(value).strip()
			if value.endswith("-"):
				category = value[:-1].strip() or "General"
			else:
				frequency = normalize_frequency(row[frequency_column - 1])
				items.append((category, value, frequency))
		return items
	finally:
		workbook.close()


def create_tasks_from_template(tasks: list[Task], list_name: str, bill_for_month: str) -> int:
	# Import only missing tasks for the selected list, month, category, and frequency.
	created_month = month_key_for_bill(bill_for_month)
	template_items = read_template_items(list_name)
	created = 0
	for category, title, frequency in template_items:
		already_exists = any(
			task_month(task) == created_month
			and task.list_name == list_name
			and task.category == category
			and task.title == title
			and normalize_frequency(task.frequency) == frequency
			for task in tasks
		)
		if not already_exists:
			tasks.append(
				Task(
					id=uuid4().hex,
					title=title,
					created_at=now(),
					month=created_month,
					bill_for_month=bill_for_month,
					list_name=list_name,
					category=category,
					frequency=frequency,
				)
			)
			created += 1
	if created:
		save_tasks(tasks)
	return created


def choose_template_list() -> Optional[str]:
	print("\nCreate tasks from TaskSubTask")
	print("1. Pune List")
	print("2. Gaya List")
	print("3. Both lists")
	choice = input("Choose a list: ").strip()
	if choice == "1":
		return "Pune List"
	if choice == "2":
		return "Gaya List"
	if choice == "3":
		return "Both"
	print("Choose 1, 2, or 3.")
	return None


def choose_bill_month() -> str:
	current_month = datetime.now().strftime("%b").upper()
	print("\nBill for the Month")
	for number, month_name in enumerate(MONTH_OPTIONS, start=1):
		print(f"{number}. {month_name}")
	choice = input(f"Choose month [{current_month}]: ").strip().upper()
	if not choice:
		return current_month
	if choice in MONTH_OPTIONS:
		return choice
	try:
		return MONTH_OPTIONS[int(choice) - 1]
	except (ValueError, IndexError):
		print("Invalid month. Using the current month.")
		return current_month


def choose_task_title() -> Optional[str]:
	print("\nTask naming templates")
	print("1. Custom task name")
	for number, template in enumerate(STANDARD_TASK_TEMPLATES, start=2):
		print(f"{number}. {template}")

	choice = input("Choose a template: ").strip()
	if choice == "1":
		return input("Task title: ").strip() or None

	try:
		return STANDARD_TASK_TEMPLATES[int(choice) - 2]
	except (ValueError, IndexError):
		print("Choose one of the listed naming templates.")
		return None


def create_next_month_task(tasks: list[Task], completed_task: Task) -> bool:
	# Monthly tasks recur after one month; quarterly tasks recur after three months.
	target_month = next_task_month(completed_task)
	if any(
		task_month(task) == target_month
		and task.title == completed_task.title
		and task.list_name == completed_task.list_name
		and task.category == completed_task.category
		for task in tasks
	):
		return False

	tasks.append(
		Task(
			id=uuid4().hex,
			title=completed_task.title,
			created_at=now(),
			month=target_month,
			bill_for_month=datetime.strptime(target_month, "%Y-%m").strftime("%b").upper(),
			list_name=completed_task.list_name,
			category=completed_task.category,
			frequency=normalize_frequency(completed_task.frequency),
		)
	)
	return True


def complete_task(tasks: list[Task], task_number: int, selected_month: str) -> None:
	complete_tasks(tasks, [task_number], selected_month)


def complete_tasks(
	tasks: list[Task],
	task_numbers: list[int],
	selected_month: str,
	completed_at: Optional[str] = None,
) -> None:
	# Resolve the displayed task numbers, mark each pending task done, and create its next instance.
	visible_tasks = [
		task for task in tasks
		if task_month(task) == selected_month and not task.is_done
	]
	visible_tasks.sort(key=lambda task: (task.is_done, task.created_at))

	selected_tasks = []
	for task_number in dict.fromkeys(task_numbers):
		if not 1 <= task_number <= len(visible_tasks):
			print(f"Task number {task_number} does not exist.")
			continue
		task = visible_tasks[task_number - 1]
		if not task.is_done:
			selected_tasks.append(task)

	if not selected_tasks:
		return

	completed_at = completed_at or now()
	created_next_month = 0
	for task in selected_tasks:
		task.completed_at = completed_at
		created_next_month += create_next_month_task(tasks, task)
	save_tasks(tasks)
	print(f"Completed {len(selected_tasks)} task(s).")
	if created_next_month:
		print(f"Created {created_next_month} next-month task instance(s).")


def complete_tasks_by_category(tasks: list[Task], category: str, selected_month: str) -> None:
	visible_tasks = [
		task for task in tasks
		if task_month(task) == selected_month and not task.is_done
	]
	visible_tasks.sort(key=lambda task: (task.is_done, task.created_at))
	task_numbers = [
		number
		for number, task in enumerate(visible_tasks, start=1)
		if (task.category or "General") == category
	]
	if not task_numbers:
		print(f"No pending tasks found in category '{category}'.")
		return
	complete_tasks(tasks, task_numbers, selected_month)


def choose_completion_category(tasks: list[Task], selected_month: str) -> Optional[str]:
	categories = sorted(
		{
			task.category or "General"
			for task in tasks
			if task_month(task) == selected_month and not task.is_done
		}
	)
	if not categories:
		print("No pending categories found for this month.")
		return None

	print("\nPending categories")
	for number, category in enumerate(categories, start=1):
		print(f"{number}. {category}")
	try:
		choice = int(input("Choose a category: "))
		return categories[choice - 1]
	except (ValueError, IndexError):
		print("Choose a valid category number.")
		return None


def choose_month(tasks: list[Task]) -> str:
	# Users choose readable MON values while the application continues using YYYY-MM internally.
	current_month = datetime.now().strftime("%Y-%m")
	available_month_keys = sorted({task_month(task) for task in tasks}, reverse=True)
	available_months = [
		datetime.strptime(value, "%Y-%m").strftime("%b").upper()
		for value in available_month_keys
	]
	print(f"\nAvailable months: {', '.join(available_months) or 'none'}")
	current_month_label = datetime.strptime(current_month, "%Y-%m").strftime("%b").upper()
	selected_month = input(f"Month [{current_month_label}]: ").strip().upper() or current_month_label
	if selected_month in MONTH_OPTIONS:
		matching_months = [
			value for value in available_month_keys
			if datetime.strptime(value, "%Y-%m").strftime("%b").upper() == selected_month
		]
		return matching_months[0] if matching_months else month_key_for_bill(selected_month)
	try:
		datetime.strptime(selected_month, "%Y-%m")
	except ValueError:
		print("Use the YYYY-MM format, for example 2026-08.")
		return current_month
	return selected_month


def complete_tasks_interactively(tasks: list[Task]) -> None:
	while True:
		available_lists = sorted({task.list_name or "Other" for task in tasks})
		print(f"\nAvailable lists: {', '.join(available_lists) or 'none'}")
		selected_month = choose_month(tasks)
		display_tasks(tasks, selected_month, pending_only=True)

		if not any(task_month(task) == selected_month and not task.is_done for task in tasks):
			pass
		else:
			try:
				print("\nComplete tasks by:")
				print("1. Category")
				print("2. Task number(s)")
				completion_mode = input("Choose an option: ").strip()
				if completion_mode == "1":
					category = choose_completion_category(tasks, selected_month)
					if category:
						complete_tasks_by_category(tasks, category, selected_month)
				elif completion_mode == "2":
					numbers = input("Task numbers to mark done (comma-separated): ")
					task_numbers = [int(number.strip()) for number in numbers.split(",") if number.strip()]
					complete_tasks(tasks, task_numbers, selected_month)
				else:
					print("Choose 1 for category or 2 for task numbers.")
			except ValueError:
				print("Enter task numbers separated by commas, for example: 1, 3, 5.")

		if not any(task_month(task) == selected_month and not task.is_done for task in tasks):
			print(f"All tasks are already Done/Completed for {selected_month}.")
			choice = input("Select another month or exit? (1 = another month, 2 = exit): ").strip()
			if choice != "1":
				return


def run() -> None:
	# The CLI keeps task creation and completion short; detailed work is handled by the functions above.
	tasks = load_tasks()
	current_month = datetime.now().strftime("%Y-%m")

	while True:
		print("Monthly Todo Tracker")
		print("1. Create tasks from Pune/Gaya list")
		print("2. Add custom task")
		print("3. Mark multiple tasks done")
		print("4. Exit")
		choice = input("Choose an option: ").strip()

		if choice == "1":
			selected_list = choose_template_list()
			bill_for_month = choose_bill_month() if selected_list else None
			if selected_list == "Both":
				created = sum(
					create_tasks_from_template(tasks, list_name, bill_for_month)
					for list_name in ("Pune List", "Gaya List")
				)
			elif selected_list:
				created = create_tasks_from_template(tasks, selected_list, bill_for_month)
			else:
				created = 0
			print(f"Created {created} new task(s) for this month.")
		elif choice == "2":
			title = choose_task_title()
			if title:
				add_task(tasks, title, choose_bill_month())
			else:
				print("Task title cannot be empty.")
		elif choice == "3":
			complete_tasks_interactively(tasks)
		elif choice == "4":
			print("Goodbye!")
			break
		else:
			print("Choose a number from 1 to 4.")


def add_template_item(list_name: str, category: str, title: str, frequency: str) -> None:
	"""Append a new category and task to the selected TaskSubTask list."""
	sync_workbook()
	workbook = load_workbook(DATA_FILE)
	try:
		if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
			workbook.create_sheet(TEMPLATE_SHEET_NAME)
		worksheet = workbook[TEMPLATE_SHEET_NAME]
		value_column = 1 if list_name == "Pune List" else 4
		frequency_column = value_column + 1
		worksheet.cell(1, value_column, list_name)
		worksheet.cell(1, frequency_column, "Bill Frequency")
		worksheet.cell(worksheet.max_row + 1, value_column, f"{category.strip()}-")
		row_number = worksheet.max_row + 1
		worksheet.cell(row_number, value_column, title.strip())
		worksheet.cell(row_number, frequency_column, normalize_frequency(frequency))
		workbook.save(DATA_FILE)
	finally:
		workbook.close()
	sync_workbook()


class TodoDesktopApp:
	"""Small desktop interface for the Excel-backed todo tracker."""

	def __init__(self, root: tk.Tk) -> None:
		self.root = root
		self.root.title("Monthly Todo Tracker")
		self.root.geometry("1050x620")
		self.tasks = load_tasks()
		self.month_keys: dict[str, str] = {}
		self.row_tasks: dict[str, Task] = {}
		self.month_var = tk.StringVar()
		self.list_var = tk.StringVar(value="Both")
		self.filter_category_var = tk.StringVar(value="All categories")
		self.frequency_var = tk.StringVar(value="All frequencies")
		self.status_var_filter = tk.StringVar(value="All statuses")
		self.status_message_var = tk.StringVar(value="Ready")
		self.title_var = tk.StringVar()
		self.numbers_var = tk.StringVar()
		self.category_var = tk.StringVar()
		self.edit_title_var = tk.StringVar()
		self.edit_list_var = tk.StringVar()
		self.edit_category_var = tk.StringVar()
		self.edit_frequency_var = tk.StringVar()
		self.edit_bill_month_var = tk.StringVar()
		self.template_list_var = tk.StringVar(value="Pune List")
		self.template_category_var = tk.StringVar()
		self.template_title_var = tk.StringVar()
		self.template_frequency_var = tk.StringVar(value="Monthly")
		self.status_var = tk.StringVar(value="Ready")
		self.month_combo: ttk.Combobox
		self.list_combo: ttk.Combobox
		self.filter_category_combo: ttk.Combobox
		self.frequency_combo: ttk.Combobox
		self.status_combo: ttk.Combobox
		self.build_widgets()
		self.refresh()

	def build_widgets(self) -> None:
		header = ttk.Frame(self.root, padding=(12, 12, 12, 0))
		header.pack(fill="x")
		tk.Label(header, text="Filter tasks", font=("TkDefaultFont", 10, "bold")).pack(side="left")
		tk.Button(header, text="Create task", command=self.open_create_task_form).pack(side="left", padx=16)

		filters = ttk.Frame(self.root, padding=10)
		filters.pack(fill="x", padx=12, pady=(4, 12))
		filter_definitions = (
			("Month", "month_combo", self.month_var, 12),
			("Category", "filter_category_combo", self.filter_category_var, 22),
			("List", "list_combo", self.list_var, 14),
			("Frequency", "frequency_combo", self.frequency_var, 14),
			("Status", "status_combo", self.status_var_filter, 14),
		)
		for column, (label, attribute, variable, width) in enumerate(filter_definitions):
			ttk.Label(filters, text=label).grid(row=0, column=column, padx=5, sticky="w")
			combo = ttk.Combobox(filters, textvariable=variable, state="readonly", width=width)
			combo.grid(row=1, column=column, padx=5, sticky="w")
			combo.bind("<<ComboboxSelected>>", self.apply_filters)
			setattr(self, attribute, combo)

		table_frame = ttk.Frame(self.root, padding=(12, 0))
		table_frame.pack(fill="both", expand=True)
		columns = ("number", "list", "category", "task", "frequency", "bill", "status")
		self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="extended")
		labels = {
			"number": "#", "list": "List", "category": "Category", "task": "Task",
			"frequency": "Frequency", "bill": "Bill for", "status": "Status",
		}
		widths = {"number": 45, "list": 110, "category": 130, "task": 350, "frequency": 95, "bill": 80, "status": 90}
		for column in columns:
			self.tree.heading(column, text=labels[column])
			self.tree.column(column, width=widths[column], anchor="w")
		self.tree.tag_configure("pending", background="white")
		self.tree.tag_configure("done", background="#C6EFCE")
		self.tree.pack(side="left", fill="both", expand=True)
		scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
		scrollbar.pack(side="right", fill="y")
		self.tree.configure(yscrollcommand=scrollbar.set)
		self.empty_filter_message = ttk.Label(
			table_frame,
			text="Select your required filter above to see the tasks",
			anchor="center",
			font=("TkDefaultFont", 12),
		)

		operation = ttk.Frame(self.root, padding=12)
		operation.pack(fill="x")
		tk.Button(operation, text="Mark selected tasks as completed", command=self.complete_selected_rows).pack(side="left", padx=(0, 8))
		tk.Button(operation, text="Save", command=self.save_from_ui).pack(side="left", padx=4)
		tk.Button(operation, text="Clear", command=self.clear_from_ui).pack(side="left", padx=4)
		tk.Button(operation, text="Exit", command=self.exit_from_ui).pack(side="left", padx=4)
		tk.Label(operation, textvariable=self.status_message_var).pack(side="left", padx=16)

	def open_create_task_form(self) -> None:
		form = tk.Toplevel(self.root)
		form.title("Create task")
		form.resizable(False, False)
		form.transient(self.root)
		form.grab_set()

		list_var = tk.StringVar(value="Pune List")
		category_var = tk.StringVar()
		title_var = tk.StringVar()
		frequency_var = tk.StringVar(value="Monthly")
		bill_month_var = tk.StringVar(value=datetime.now().strftime("%b").upper())

		fields = ttk.Frame(form, padding=16)
		fields.pack(fill="both", expand=True)
		category_options = sorted({
			(task.category or "General")
			for task in self.tasks
		} | {
			category
			for list_name in ("Pune List", "Gaya List")
			for category, _title, _frequency in read_template_items(list_name)
		})
		if category_options:
			category_var.set(category_options[0])

		controls = (
			("List", ttk.Combobox(fields, textvariable=list_var, values=("Pune List", "Gaya List"), state="readonly", width=24)),
			("Category", ttk.Combobox(fields, textvariable=category_var, values=category_options, state="readonly", width=24)),
			("Task", ttk.Entry(fields, textvariable=title_var, width=27)),
			("Frequency", ttk.Combobox(fields, textvariable=frequency_var, values=("Monthly", "Quarterly"), state="readonly", width=24)),
			("Bill for Month", ttk.Combobox(fields, textvariable=bill_month_var, values=MONTH_OPTIONS, state="readonly", width=24)),
		)
		for row, (label, control) in enumerate(controls):
			ttk.Label(fields, text=label).grid(row=row, column=0, padx=(0, 10), pady=5, sticky="w")
			control.grid(row=row, column=1, pady=5, sticky="ew")

		ttk.Label(fields, text="Status: Pending").grid(row=5, column=0, columnspan=2, pady=(8, 2), sticky="w")
		tk.Label(fields, text="Created Date: automatic    Completed Date: blank").grid(row=6, column=0, columnspan=2, pady=2, sticky="w")

		def create() -> None:
			category = category_var.get().strip()
			title = title_var.get().strip()
			bill_for_month = bill_month_var.get().strip().upper()
			if not category or not title or bill_for_month not in MONTH_OPTIONS:
				messagebox.showwarning("Task details", "List, Category, Task, Frequency, and Bill for Month are required.", parent=form)
				return
			self.tasks.append(
				Task(
					id=uuid4().hex,
					title=title,
					created_at=now(),
					month=month_key_for_bill(bill_for_month),
					bill_for_month=bill_for_month,
					list_name=list_var.get(),
					category=category,
					frequency=normalize_frequency(frequency_var.get()),
				)
			)
			save_tasks(self.tasks)
			form.destroy()
			self.refresh()
			self.status_message_var.set(f"Created task: {title}")

		ttk.Button(fields, text="Create", command=create).grid(row=7, column=0, columnspan=2, pady=(12, 0))

	def save_from_ui(self) -> None:
		save_tasks(self.tasks)
		self.status_message_var.set("Saved to both MyToDoList.xlsx locations.")

	def clear_from_ui(self) -> None:
		self.month_var.set("")
		self.list_var.set("")
		self.filter_category_var.set("")
		self.frequency_var.set("")
		self.status_var_filter.set("")
		self.tree.selection_set(())
		self.update_filter_categories()
		self.populate_tasks()
		self.status_message_var.set("Filters and selection cleared.")

	def exit_from_ui(self) -> None:
		save_tasks(self.tasks)
		self.root.destroy()

	def selected_month_key(self) -> str:
		return self.month_keys.get(self.month_var.get(), datetime.now().strftime("%Y-%m"))

	def refresh(self) -> None:
		self.tasks = load_tasks()
		month_values = sorted({task_month(task) for task in self.tasks} | {datetime.now().strftime("%Y-%m")}, reverse=True)
		self.month_keys = {
			datetime.strptime(value, "%Y-%m").strftime("%b").upper(): value for value in month_values
		}
		self.month_combo["values"] = ["", *self.month_keys]
		current_label = datetime.now().strftime("%b").upper()
		self.month_var.set(current_label if current_label in self.month_keys else next(iter(self.month_keys)))
		self.list_combo["values"] = ("", "Both", "Pune List", "Gaya List")
		self.frequency_combo["values"] = ("", "All frequencies", "Monthly", "Quarterly")
		self.status_combo["values"] = ("", "All statuses", PENDING_STATUS, DONE_STATUS)
		if self.frequency_var.get() not in self.frequency_combo["values"]:
			self.frequency_var.set("All frequencies")
		if self.status_var_filter.get() not in self.status_combo["values"]:
			self.status_var_filter.set("All statuses")
		self.update_filter_categories()
		self.populate_tasks()

	def reset_filters(self) -> None:
		self.list_var.set("Both")
		self.filter_category_var.set("All categories")
		self.refresh()

	def apply_filters(self, _event: object = None) -> None:
		self.update_filter_categories()
		self.populate_tasks()

	def update_filter_categories(self) -> None:
		selected_month = self.selected_month_key()
		selected_list = self.list_var.get()
		selected_frequency = self.frequency_var.get()
		categories = sorted({
			task.category or "General"
			for task in self.tasks
			if task_month(task) == selected_month
			and (not selected_list or selected_list == "Both" or task.list_name == selected_list)
			and (not selected_frequency or selected_frequency == "All frequencies" or normalize_frequency(task.frequency) == selected_frequency)
		})
		values = ["", "All categories", *categories]
		self.filter_category_combo["values"] = values
		if self.filter_category_var.get() not in values:
			self.filter_category_var.set("")

	def pending_tasks(self) -> list[Task]:
		selected_month = self.selected_month_key()
		selected_list = self.list_var.get()
		selected_category = self.filter_category_var.get()
		selected_frequency = self.frequency_var.get()
		return [
			task for task in self.tasks
			if task_month(task) == selected_month
			and not task.is_done
			and (not selected_list or selected_list == "Both" or task.list_name == selected_list)
			and (not selected_category or selected_category == "All categories" or (task.category or "General") == selected_category)
			and (not selected_frequency or selected_frequency == "All frequencies" or normalize_frequency(task.frequency) == selected_frequency)
			and (not self.status_var_filter.get() or self.status_var_filter.get() in ("All statuses", PENDING_STATUS))
		]

	def populate_tasks(self) -> None:
		for item in self.tree.get_children():
			self.tree.delete(item)
		filters_are_clear = not any((
			self.month_var.get(),
			self.filter_category_var.get(),
			self.list_var.get(),
			self.frequency_var.get(),
			self.status_var_filter.get(),
		))
		if filters_are_clear:
			self.empty_filter_message.place(relx=0.5, rely=0.5, anchor="center")
			self.status_message_var.set("Select a filter to display tasks.")
			self.row_tasks.clear()
			return
		self.empty_filter_message.place_forget()
		selected_month = self.selected_month_key()
		selected_list = self.list_var.get()
		selected_category = self.filter_category_var.get()
		selected_frequency = self.frequency_var.get()
		selected_status = self.status_var_filter.get()
		visible = [
			task for task in self.tasks
			if task_month(task) == selected_month
			and (not selected_list or selected_list == "Both" or task.list_name == selected_list)
			and (not selected_category or selected_category == "All categories" or (task.category or "General") == selected_category)
			and (not selected_frequency or selected_frequency == "All frequencies" or normalize_frequency(task.frequency) == selected_frequency)
			and (not selected_status or selected_status == "All statuses" or (DONE_STATUS if task.is_done else PENDING_STATUS) == selected_status)
		]
		self.row_tasks.clear()
		for number, task in enumerate(visible, start=1):
			iid = self.tree.insert("", "end", values=(number, task.list_name or "", task.category or "General", task.title, normalize_frequency(task.frequency), bill_month(task), DONE_STATUS if task.is_done else PENDING_STATUS), tags=("done" if task.is_done else "pending",))
			self.row_tasks[iid] = task
		self.status_message_var.set(f"Showing {len(visible)} task(s) for {self.month_var.get()}")

	def create_template_tasks(self) -> None:
		month = self.month_var.get()
		selected_list = self.list_var.get()
		lists = ("Pune List", "Gaya List") if selected_list == "Both" else (selected_list,)
		created = sum(create_tasks_from_template(self.tasks, list_name, month) for list_name in lists)
		self.refresh()
		self.status_var.set(f"Created {created} task(s) for {month}.")

	def add_custom_task(self) -> None:
		title = self.title_var.get().strip()
		if not title:
			messagebox.showwarning("Task title", "Enter a task title first.")
			return
		add_task(self.tasks, title, self.month_var.get())
		self.title_var.set("")
		self.refresh()

	def load_selected_task(self, _event: object = None) -> None:
		selection = self.tree.selection()
		if len(selection) != 1:
			return
		task = self.row_tasks[selection[0]]
		if task.is_done:
			self.clear_edit_fields()
			return
		self.edit_title_var.set(task.title)
		self.edit_list_var.set(task.list_name or "")
		self.edit_category_var.set(task.category or "")
		self.edit_frequency_var.set(normalize_frequency(task.frequency))
		self.edit_bill_month_var.set(bill_month(task))

	def clear_edit_fields(self) -> None:
		for variable in (self.edit_title_var, self.edit_list_var, self.edit_category_var, self.edit_frequency_var, self.edit_bill_month_var):
			variable.set("")

	def update_selected_task(self) -> None:
		selection = self.tree.selection()
		if len(selection) != 1:
			messagebox.showinfo("Select one task", "Select one pending task to edit.")
			return
		task = self.row_tasks[selection[0]]
		if task.is_done:
			messagebox.showinfo("Completed task", "Completed tasks cannot be modified.")
			return
		if not self.edit_title_var.get().strip() or not self.edit_bill_month_var.get():
			messagebox.showwarning("Task details", "Task title and billing month are required.")
			return
		task.title = self.edit_title_var.get().strip()
		task.list_name = self.edit_list_var.get() or None
		task.category = self.edit_category_var.get().strip() or None
		task.frequency = normalize_frequency(self.edit_frequency_var.get())
		task.bill_for_month = self.edit_bill_month_var.get()
		task.month = month_key_for_bill(task.bill_for_month)
		save_tasks(self.tasks)
		self.refresh()

	def add_new_template(self) -> None:
		category = self.template_category_var.get().strip()
		title = self.template_title_var.get().strip()
		if not category or not title:
			messagebox.showwarning("Template task", "Category and task title are required.")
			return
		add_template_item(self.template_list_var.get(), category, title, self.template_frequency_var.get())
		self.template_category_var.set("")
		self.template_title_var.set("")
		self.tasks = load_tasks()
		messagebox.showinfo("Template task", "Task added to TaskSubTask.")

	def complete_by_number(self) -> None:
		try:
			numbers = [int(value.strip()) for value in self.numbers_var.get().split(",") if value.strip()]
		except ValueError:
			messagebox.showwarning("Task numbers", "Use comma-separated numbers such as 1, 3, 5.")
			return
		pending = self.pending_tasks()
		selected_tasks = [pending[number - 1] for number in dict.fromkeys(numbers) if 1 <= number <= len(pending)]
		self.complete_selected(selected_tasks)

	def complete_by_category(self) -> None:
		category = self.category_var.get()
		self.complete_selected([task for task in self.pending_tasks() if (task.category or "General") == category])

	def complete_selected_rows(self) -> None:
		selected_tasks = [self.row_tasks[item] for item in self.tree.selection() if not self.row_tasks[item].is_done]
		if not selected_tasks:
			messagebox.showinfo("No pending tasks", "Select one or more pending tasks.")
			return
		completion_date = simpledialog.askstring(
			"Completion date",
			"Enter completion date (YYYY-MMM-DD, for example 2026-AUG-20):",
			initialvalue=datetime.now().strftime(COMPLETION_DATE_FORMAT).upper(),
			parent=self.root,
		)
		if completion_date is None:
			return
		try:
			completed_date = datetime.strptime(completion_date.strip().upper(), COMPLETION_DATE_FORMAT)
		except ValueError:
			messagebox.showwarning("Invalid date", f"Use this format: {COMPLETION_DATE_FORMAT.upper()} (example: 2026-AUG-20)")
			return
		pending = self.pending_tasks()
		selected_ids = {task.id for task in selected_tasks}
		numbers = [number for number, task in enumerate(pending, start=1) if task.id in selected_ids]
		complete_tasks(self.tasks, numbers, self.selected_month_key(), completed_date.strftime(COMPLETION_DATE_FORMAT).upper())
		self.refresh()

	def complete_selected(self, selected_tasks: list[Task]) -> None:
		if not selected_tasks:
			messagebox.showinfo("No tasks", "No pending tasks were selected.")
			return
		selected_ids = {task.id for task in selected_tasks}
		complete_tasks(self.tasks, [number for number, task in enumerate(self.pending_tasks(), start=1) if task.id in selected_ids], self.selected_month_key())
		self.numbers_var.set("")
		self.refresh()


def run_gui() -> None:
	root = tk.Tk()
	TodoDesktopApp(root)
	root.mainloop()


if __name__ == "__main__":
	run_gui()
