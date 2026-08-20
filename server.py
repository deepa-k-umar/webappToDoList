from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).parent
WORKBOOK_PATH = ROOT / "MyToDoList.xlsx"
HEADERS = ("ID", "List", "Category", "Task", "Frequency", "Bill for the Month", "Month Key", "Created Date", "Completed Date", "Status")


def append_tasks(tasks):
    if WORKBOOK_PATH.exists():
        workbook = load_workbook(WORKBOOK_PATH)
        if "Todo List" in workbook.sheetnames:
            sheet = workbook["Todo List"]
        else:
            sheet = workbook.create_sheet("Todo List", 0)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Todo List"

    if sheet.max_row == 1 and not sheet.cell(1, 1).value:
        sheet.delete_rows(1)
    if sheet.max_row == 0 or not sheet.cell(1, 1).value:
        sheet.append(HEADERS)

    headers = [cell.value for cell in sheet[1]]
    if headers != list(HEADERS):
        sheet.delete_rows(1)
        sheet.insert_rows(1)
        for column, header in enumerate(HEADERS, 1):
            sheet.cell(1, column).value = header
        headers = list(HEADERS)

    id_column = headers.index("ID") + 1
    existing_ids = {str(sheet.cell(row, id_column).value) for row in range(2, sheet.max_row + 1) if sheet.cell(row, id_column).value}
    created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    added = 0
    for task in tasks:
        task_id = str(task.get("id", "")).strip()
        title = str(task.get("title", "")).strip()
        if not task_id or not title or task_id in existing_ids:
            continue
        month = str(task.get("month", ""))
        bill_month = datetime.strptime(month, "%Y-%m").strftime("%b").upper() if month else ""
        completed = created if task.get("done") else ""
        sheet.append([task_id, "", "", title, task.get("frequency", "Monthly"), bill_month, month, created, completed, "Done" if task.get("done") else "Pending"])
        existing_ids.add(task_id)
        added += 1

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(WORKBOOK_PATH)
    workbook.close()
    return added


class Handler(SimpleHTTPRequestHandler):
    def do_POST(self):
        if self.path != "/export-excel":
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", 0))
            tasks = json.loads(self.rfile.read(length))
            added = append_tasks(tasks)
            response = json.dumps({"added": added}).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(response)))
            self.end_headers()
            self.wfile.write(response)
        except (OSError, ValueError, TypeError, json.JSONDecodeError) as error:
            response = json.dumps({"error": str(error)}).encode()
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(response)))
            self.end_headers()
            self.wfile.write(response)

    def do_GET(self):
        super().do_GET()


if __name__ == "__main__":
    print(f"Serving Monthly Todo at http://localhost:8000")
    print(f"Excel file: {WORKBOOK_PATH}")
    ThreadingHTTPServer(("localhost", 8000), lambda *args, **kwargs: Handler(*args, directory=str(ROOT), **kwargs)).serve_forever()
